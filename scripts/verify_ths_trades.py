"""验证同花顺交易文件一致性

检查 output/ths/industry_rotation_v5/ 下所有 xlsx 文件:
1. 现金流: 现金始终 >= 0
2. 持仓: 数量为 100 的整数倍
3. 费用: 成交金额 * 0.001
4. 持仓数: 累计买入 - 累计卖出 = 实际持仓
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

CAPITAL = 10_000_000
COST_RATE = 0.001
THS_DIR = Path("output/ths/industry_rotation_v5")

EXPECTED_REBALANCE_DATES = [
    "2026-06-01",  # build
    "2026-06-08",
    "2026-06-15",
    "2026-06-22",
    "2026-06-29",
    "2026-07-06",
    "2026-07-13",
]


def load_trades(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Sheet1"]
    headers = ["交易日期", "证券代码", "业务类型", "数量", "价格",
               "成交金额", "费用", "证券类型", "说明"]
    trades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        trades.append(dict(zip(headers, row)))
    wb.close()
    return trades


def main():
    files = sorted(THS_DIR.glob("*.xlsx"))
    print(f"找到 {len(files)} 个文件:")
    for f in files:
        print(f"  {f.name}")
    print()

    cash = CAPITAL
    holdings: dict[str, int] = {}
    errors: list[str] = []
    warnings: list[str] = []

    for f in files:
        print(f"=== {f.name} ===")
        trades = load_trades(f)
        if not trades:
            print("  (空文件)")
            continue

        file_cash_delta = 0.0
        for t in trades:
            dt = t["交易日期"]
            code = t["证券代码"]
            biz = t["业务类型"]
            qty = t["数量"]
            price = t["价格"]
            amount = t["成交金额"]
            cost = t["费用"]

            # 1. 数量为 100 整数倍
            if qty % 100 != 0:
                errors.append(f"{f.name}: {code} 数量 {qty} 不是 100 整数倍")

            # 2. 费用 = 成交金额 * 0.001
            expected_cost = round(amount * COST_RATE, 2)
            if abs(cost - expected_cost) > 0.01:
                errors.append(f"{f.name}: {code} 费用 {cost} != 预期 {expected_cost}")

            # 3. 成交金额 = 数量 * 价格
            expected_amount = round(qty * price, 2)
            if abs(amount - expected_amount) > 0.01:
                errors.append(f"{f.name}: {code} 成交金额 {amount} != 预期 {expected_amount}")

            # 4. 更新持仓与现金
            if biz == "买入":
                holdings[code] = holdings.get(code, 0) + qty
                cash -= amount + cost
                file_cash_delta -= amount + cost
            elif biz == "卖出":
                if holdings.get(code, 0) < qty:
                    errors.append(f"{f.name}: {code} 卖出 {qty} 超过持仓 {holdings.get(code, 0)}")
                holdings[code] = holdings.get(code, 0) - qty
                if holdings[code] == 0:
                    holdings.pop(code)
                cash += amount - cost
                file_cash_delta += amount - cost

            action = "买" if biz == "买入" else "卖"
            print(f"  {dt.strftime('%Y-%m-%d')} {action} {code:<12} {qty:>6d}@{price:.2f} 金额={amount:.0f} 费用={cost:.0f}")

        print(f"  -> 文件现金流: {file_cash_delta:+,.0f}")
        print(f"  -> 累计现金: {cash:,.0f}")
        print(f"  -> 累计持仓: {len(holdings)} 只")

        # 5. 现金非负
        if cash < 0:
            errors.append(f"{f.name}: 现金为负 {cash:.0f}")
        print()

    print("=" * 60)
    print(f"最终持仓: {len(holdings)} 只  现金: {cash:,.0f}")
    print(f"总资产: {cash + sum(holdings.values()) * 0:.0f} (按0价计算, 仅参考)")

    print(f"\n持仓明细 (按数量):")
    for code, qty in sorted(holdings.items(), key=lambda x: -x[1]):
        print(f"  {code:<12} {qty:>6d} 股")

    print(f"\n{'='*60}")
    if errors:
        print(f"ERRORS ({len(errors)}):")
        for e in errors:
            print(f"  [ERR] {e}")
    else:
        print("无 ERROR")

    if warnings:
        print(f"\nWARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  [WARN] {w}")

    # 与期望调仓日对比
    actual_dates = [f.stem[:4] + "-" + f.stem[4:6] + "-" + f.stem[6:8] for f in files]
    print(f"\n期望调仓日: {EXPECTED_REBALANCE_DATES}")
    print(f"实际文件日: {actual_dates}")
    missing = set(EXPECTED_REBALANCE_DATES) - set(actual_dates)
    extra = set(actual_dates) - set(EXPECTED_REBALANCE_DATES)
    if missing:
        print(f"缺失文件: {sorted(missing)}")
    if extra:
        print(f"额外文件: {sorted(extra)}")

    return 0 if not errors else 1


if __name__ == "__main__":
    sys.exit(main())
