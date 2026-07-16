"""同花顺模拟盘交易流水导出（行业轮动策略 YCJ_industry_v5）

从 OOS 回测结果生成同花顺 PMS 交易组合流水导入文件。
每次调仓生成一个独立 xlsx 文件，便于逐日上传同花顺模拟盘。

文件组织:
  templates/ths_pms_template.xlsx                    # 同花顺模板
  output/ths/YCJ_industry_v5/
    20260601_build.xlsx                              # 建仓文件
    20260608_rebalance.xlsx                          # 调仓文件
    ...

流水格式（同花顺模板）:
  交易日期 | 证券代码 | 业务类型 | 数量 | 价格 | 成交金额 | 费用 | 证券类型 | 说明

用法:
    # 生成全部 OOS 建仓+调仓文件（0601~0715）
    python scripts/ind_ths_export.py

    # 指定版本和初始资金
    python scripts/ind_ths_export.py --version v5 --capital 10000000

    # 只生成指定日期的调仓文件（日常每日运行用）
    python scripts/ind_ths_export.py --date 20260715
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook, load_workbook

from ohmyquant.data.sources.duckdb_source import DuckDBSource

STRATEGY_NAME = "YCJ_industry_v5"
DATA_ROOT = "D:/Work/Project/download_a_share/data"
TEMPLATE_PATH = Path("templates/ths_pms_template.xlsx")
OUTPUT_DIR = Path(f"output/ths/{STRATEGY_NAME}")
TRANSACTION_COST_RATE = 0.001  # 千分之一交易费
LOT_SIZE = 100  # A股最小交易单位（1手=100股）


def load_oos_result(version: str) -> dict:
    """加载 OOS 回测结果"""
    oos_file = Path(f"output/oos_compare/ind/{version}_oos.json")
    if not oos_file.exists():
        raise FileNotFoundError(
            f"OOS 回测结果不存在: {oos_file}\n"
            f"请先运行: python scripts/ind_oos.py {version}"
        )
    with open(oos_file, "r", encoding="utf-8") as f:
        return json.load(f)


def get_open_prices(source: DuckDBSource, codes: list[str], date_str: str) -> dict[str, float]:
    """获取指定日期的开盘价"""
    df = source.load_daily_price(codes, date_str, date_str, adjust="post")
    if df is None or len(df) == 0:
        # 尝试不复权
        df = source.load_daily_price(codes, date_str, date_str, adjust="none")
    if df is None or len(df) == 0:
        return {}

    result: dict[str, float] = {}
    for row in df.iter_rows(named=True):
        code = row.get("code", "")
        open_price = row.get("open")
        if open_price and isinstance(open_price, (int, float)):
            result[code] = float(open_price)
    return result


def compute_lot_shares(capital: float, weight: float, price: float) -> int:
    """计算股数（取整到100股）"""
    if price <= 0:
        return 0
    target_value = capital * weight
    raw_shares = target_value / price
    return int(raw_shares // LOT_SIZE) * LOT_SIZE


def generate_build_trades(
    date_str: str,
    holdings: dict[str, float],
    open_prices: dict[str, float],
    capital: float,
) -> list[dict]:
    """生成建仓交易流水（全部买入）"""
    trades: list[dict] = []
    for code, weight in holdings.items():
        price = open_prices.get(code)
        if not price or price <= 0:
            print(f"  [警告] {code} 无开盘价，跳过")
            continue
        shares = compute_lot_shares(capital, weight, price)
        if shares <= 0:
            continue
        amount = shares * price
        cost = amount * TRANSACTION_COST_RATE
        trades.append({
            "交易日期": datetime.strptime(date_str, "%Y-%m-%d"),
            "证券代码": code,
            "业务类型": "买入",
            "数量": shares,
            "价格": round(price, 4),
            "成交金额": round(amount, 2),
            "费用": round(cost, 2),
            "证券类型": "A股",
            "说明": f"建仓 {STRATEGY_NAME}",
        })
    return trades


def generate_rebalance_trades(
    date_str: str,
    prev_holdings_shares: dict[str, int],
    target_holdings: dict[str, float],
    open_prices: dict[str, float],
    prev_cash: float,
) -> tuple[list[dict], dict[str, int], float]:
    """生成调仓交易流水

    Returns:
        (trades, new_holdings_shares, new_cash)
    """
    # 计算当前总资产 = 持仓市值 + 现金
    current_value = prev_cash
    for code, shares in prev_holdings_shares.items():
        price = open_prices.get(code, 0)
        current_value += shares * price

    # 计算目标股数
    target_shares: dict[str, int] = {}
    for code, weight in target_holdings.items():
        price = open_prices.get(code)
        if not price or price <= 0:
            print(f"  [警告] {code} 无开盘价，跳过买入")
            continue
        target_shares[code] = compute_lot_shares(current_value, weight, price)

    # 生成交易：先卖出不在目标中的 + 卖出超量，再买入
    trades: list[dict] = []
    dt = datetime.strptime(date_str, "%Y-%m-%d")

    # 卖出：不在目标持仓中 或 目标数量小于当前
    for code, shares in prev_holdings_shares.items():
        price = open_prices.get(code, 0)
        if price <= 0:
            continue
        target = target_shares.get(code, 0)
        if shares > target:
            sell_shares = shares - target
            amount = sell_shares * price
            cost = amount * TRANSACTION_COST_RATE
            trades.append({
                "交易日期": dt,
                "证券代码": code,
                "业务类型": "卖出",
                "数量": sell_shares,
                "价格": round(price, 4),
                "成交金额": round(amount, 2),
                "费用": round(cost, 2),
                "证券类型": "A股",
                "说明": f"调仓卖出 {STRATEGY_NAME}",
            })

    # 买入：目标数量大于当前
    for code, target in target_shares.items():
        price = open_prices.get(code, 0)
        if price <= 0:
            continue
        current = prev_holdings_shares.get(code, 0)
        if target > current:
            buy_shares = target - current
            amount = buy_shares * price
            cost = amount * TRANSACTION_COST_RATE
            trades.append({
                "交易日期": dt,
                "证券代码": code,
                "业务类型": "买入",
                "数量": buy_shares,
                "价格": round(price, 4),
                "成交金额": round(amount, 2),
                "费用": round(cost, 2),
                "证券类型": "A股",
                "说明": f"调仓买入 {STRATEGY_NAME}",
            })

    # 计算调仓后现金
    new_cash = prev_cash
    for t in trades:
        if t["业务类型"] == "买入":
            new_cash -= t["成交金额"] + t["费用"]
        else:
            new_cash += t["成交金额"] - t["费用"]

    return trades, target_shares, new_cash


def write_xlsx(trades: list[dict], output_path: Path):
    """基于模板写入交易流水"""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Sheet1"]

    # 清空模板示例数据（保留表头）
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None

    # 写入交易
    headers = [
        "交易日期", "证券代码", "业务类型", "数量", "价格",
        "成交金额", "费用", "证券类型", "说明",
    ]
    for i, trade in enumerate(trades, start=2):
        for j, key in enumerate(headers, start=1):
            ws.cell(i, j).value = trade[key]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def export_all(version: str, capital: float, only_date: str | None = None):
    """导出全部建仓+调仓文件"""
    oos_data = load_oos_result(version)
    rebalance_log = oos_data["rebalance_log"]

    if only_date:
        rebalance_log = [r for r in rebalance_log if r["date"] == only_date]
        if not rebalance_log:
            print(f"日期 {only_date} 无调仓记录")
            return

    source = DuckDBSource({"data_root": DATA_ROOT})
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    prev_shares: dict[str, int] = {}
    prev_cash = capital
    is_first = True

    print(f"策略: {STRATEGY_NAME}")
    print(f"初始资金: {capital:,.0f}")
    print(f"OOS调仓次数: {len(rebalance_log)}")
    print(f"输出目录: {OUTPUT_DIR}")
    print("=" * 60)

    for entry in rebalance_log:
        date_str = entry["date"]
        holdings = entry["holdings"]
        codes = list(holdings.keys())

        # 获取开盘价
        open_prices = get_open_prices(source, codes, date_str)
        if not open_prices:
            print(f"[{date_str}] 无法获取开盘价，跳过")
            continue

        if is_first:
            # 建仓
            trades = generate_build_trades(date_str, holdings, open_prices, capital)
            # 更新持仓和现金
            for t in trades:
                code = t["证券代码"]
                shares = t["数量"]
                prev_shares[code] = prev_shares.get(code, 0) + shares
                prev_cash -= t["成交金额"] + t["费用"]
            filename = f"{date_str.replace('-', '')}_build.xlsx"
            label = "建仓"
        else:
            # 调仓
            trades, prev_shares, prev_cash = generate_rebalance_trades(
                date_str, prev_shares, holdings, open_prices, prev_cash
            )
            if not trades:
                print(f"[{date_str}] 无调仓交易，跳过文件生成")
                continue
            filename = f"{date_str.replace('-', '')}_rebalance.xlsx"
            label = "调仓"

        output_path = OUTPUT_DIR / filename
        write_xlsx(trades, output_path)

        # 打印摘要
        buy_count = sum(1 for t in trades if t["业务类型"] == "买入")
        sell_count = sum(1 for t in trades if t["业务类型"] == "卖出")
        total_amount = sum(t["成交金额"] for t in trades)
        total_cost = sum(t["费用"] for t in trades)
        print(f"[{date_str}] {label}: {len(trades)}笔 (买{buy_count}/卖{sell_count}), "
              f"金额 {total_amount:,.0f}, 费用 {total_cost:,.0f}")
        for t in trades:
            action = "买" if t["业务类型"] == "买入" else "卖"
            print(f"  {action} {t['证券代码']:<12} {t['数量']:>6d}股 @{t['价格']:.2f}")

        print(f"  → {output_path}")
        print(f"  剩余现金: {prev_cash:,.0f}")
        print()

        is_first = False

    print("=" * 60)
    print("导出完成")


def main():
    parser = argparse.ArgumentParser(description="同花顺交易流水导出")
    parser.add_argument("--version", default="v5", help="策略版本")
    parser.add_argument("--capital", type=float, default=10_000_000, help="初始资金")
    parser.add_argument("--date", default=None, help="只生成指定日期(YYYY-MM-DD)的文件")
    args = parser.parse_args()

    export_all(args.version, args.capital, args.date)


if __name__ == "__main__":
    main()
