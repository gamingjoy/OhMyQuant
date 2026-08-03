"""行业轮动策略 OOS 净值分析

基于同花顺实际交易文件回放建仓与调仓，
计算 2026-06-01 以来的净值表现，对比沪深300，并给出优化建议。

用法:
    python scripts/industry_rotation_nav_analysis.py              # 默认 v53
    python scripts/industry_rotation_nav_analysis.py --version v63
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import numpy as np
import polars as pl
from openpyxl import load_workbook

from ohmyquant.data.sources.duckdb_source import DuckDBSource
import os

CAPITAL = 10_000_000
COST_RATE = 0.001
BENCHMARK = "000300.XSHG"  # 沪深300
DATA_ROOT = os.getenv("DATA_ROOT", "data")
OOS_START = "2026-06-01"


def load_trades(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Sheet1"]
    headers = ["交易日期", "证券代码", "业务类型", "数量", "价格",
               "成交金额", "费用", "证券类型", "说明"]
    trades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        trades.append(dict(zip(headers, row)))
    wb.close()
    return trades


def replay_all_trades(files: list[Path]) -> tuple[dict[str, dict[str, int]], dict[str, float], list[str]]:
    """回放所有交易文件，返回每个调仓日后的持仓与现金

    Returns:
        holdings_by_date: {date_str: {code: shares}}
        cash_by_date: {date_str: cash}
        trade_dates: 按时间排序的调仓日列表
    """
    cash = CAPITAL
    holdings: dict[str, int] = {}
    holdings_by_date: dict[str, dict[str, int]] = {}
    cash_by_date: dict[str, float] = {}
    trade_dates: list[str] = []

    for f in files:
        trades = load_trades(f)
        if not trades:
            continue
        # 文件内所有交易同一天
        date_str = trades[0]["交易日期"].strftime("%Y-%m-%d")

        for t in trades:
            code = t["证券代码"]
            qty = t["数量"]
            amount = t["成交金额"]
            cost = t["费用"]
            if t["业务类型"] == "买入":
                holdings[code] = holdings.get(code, 0) + qty
                cash -= amount + cost
            else:  # 卖出
                holdings[code] = holdings.get(code, 0) - qty
                if holdings[code] <= 0:
                    holdings.pop(code, None)
                cash += amount - cost

        holdings_by_date[date_str] = dict(holdings)
        cash_by_date[date_str] = cash
        trade_dates.append(date_str)

    return holdings_by_date, cash_by_date, trade_dates


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--version", default="v53", help="策略版本(默认 v53)")
    args = parser.parse_args()
    version = args.version

    ths_dir = Path(f"output/ths/industry_rotation_{version}")

    # 1. 加载交易文件
    files = sorted(ths_dir.glob("*.xlsx"))
    if not files:
        print(f"未找到交易文件: {ths_dir}")
        return

    print(f"加载 {len(files)} 个交易文件:")
    for f in files:
        print(f"  {f.name}")

    holdings_by_date, cash_by_date, trade_dates = replay_all_trades(files)
    print(f"\n调仓日: {trade_dates}")

    # 2. 收集所有出现过的股票代码
    all_codes: set[str] = set()
    for h in holdings_by_date.values():
        all_codes.update(h.keys())
    all_codes = sorted(all_codes)
    print(f"历史持仓股票数: {len(all_codes)}")

    # 3. 获取交易日历（OOS_START 到最新数据日）
    source = DuckDBSource({"data_root": DATA_ROOT})
    latest_date = source.get_latest_date()
    print(f"最新数据日: {latest_date}")

    trade_cal = source.get_trade_calendar(OOS_START, latest_date)
    print(f"交易日数: {len(trade_cal)}（{trade_cal[0]} → {trade_cal[-1]}）")

    # 4. 加载所有股票的实际价格日线（非复权，与同花顺成交价一致）
    price_df = source.load_daily_price(all_codes, OOS_START, latest_date, adjust="none")
    if price_df is None or len(price_df) == 0:
        print("无法加载股票价格数据")
        return

    # 构造 (date, code) -> close 价格查找表
    price_pivot = price_df.select(["date", "code", "close"]).with_columns(
        pl.col("date").dt.strftime("%Y-%m-%d")
    ).pivot(values="close", index="date", on="code", aggregate_function="first")

    # 5. 加载沪深300指数数据
    bench_df = source.load_index_data(BENCHMARK, OOS_START, latest_date)
    if bench_df is None or len(bench_df) == 0:
        print("无法加载沪深300数据")
        return
    bench_df = bench_df.with_columns(pl.col("date").dt.strftime("%Y-%m-%d")).sort("date")
    bench_close = dict(zip(bench_df["date"].to_list(), bench_df["close"].to_list()))
    bench_first = bench_df["close"][0]

    # 6. 按交易日计算净值
    # 当前生效的持仓（在两个调仓日之间保持不变）
    current_holdings: dict[str, int] = {}
    current_cash: float = CAPITAL
    rebalance_idx = 0

    records = []
    for date_str in trade_cal:
        # 如果今天是调仓日，更新持仓状态
        if rebalance_idx < len(trade_dates) and date_str == trade_dates[rebalance_idx]:
            current_holdings = holdings_by_date[date_str]
            current_cash = cash_by_date[date_str]
            rebalance_idx += 1

        # 计算当日持仓市值（用当日收盘价）
        market_value = 0.0
        if date_str in price_pivot["date"].to_list():
            row = price_pivot.filter(pl.col("date") == date_str)
            if len(row) > 0:
                row_dict = row.to_dicts()[0]
                for code, shares in current_holdings.items():
                    close = row_dict.get(code)
                    if close is not None and not np.isnan(close):
                        market_value += shares * close

        total_asset = market_value + current_cash
        nav = total_asset / CAPITAL

        bench_price = bench_close.get(date_str)
        bench_nav = bench_price / bench_first if bench_price else None

        records.append({
            "date": date_str,
            "nav": nav,
            "bench_nav": bench_nav,
            "cash": current_cash,
            "market_value": market_value,
            "n_holdings": len(current_holdings),
        })

    nav_df = pl.DataFrame(records)

    # 7. 计算绩效指标
    nav_arr = nav_df["nav"].to_numpy()
    bench_arr = nav_df["bench_nav"].to_numpy()
    n_days = len(nav_arr)

    # 日收益率
    daily_ret = np.diff(nav_arr)
    bench_ret = np.diff(bench_arr)

    # 超额日收益
    excess_ret = daily_ret - bench_ret

    # 总收益
    total_return = nav_arr[-1] - 1.0
    bench_total_return = bench_arr[-1] - 1.0
    excess_total = total_return - bench_total_return

    # 年化 Sharpe（无风险利率 2%）
    rf_daily = (1.02 ** (1 / 242)) - 1
    if len(daily_ret) > 1 and np.std(daily_ret, ddof=1) > 0:
        sharpe = (np.mean(daily_ret) - rf_daily) / np.std(daily_ret, ddof=1) * np.sqrt(242)
    else:
        sharpe = 0.0
    if len(bench_ret) > 1 and np.std(bench_ret, ddof=1) > 0:
        bench_sharpe = (np.mean(bench_ret) - rf_daily) / np.std(bench_ret, ddof=1) * np.sqrt(242)
    else:
        bench_sharpe = 0.0

    # 最大回撤
    peak = np.maximum.accumulate(nav_arr)
    drawdown = (nav_arr - peak) / peak
    max_drawdown = float(np.min(drawdown))
    bench_peak = np.maximum.accumulate(bench_arr)
    bench_dd = (bench_arr - bench_peak) / bench_peak
    bench_max_dd = float(np.min(bench_dd))

    # 年化波动率
    ann_vol = np.std(daily_ret, ddof=1) * np.sqrt(242) if len(daily_ret) > 1 else 0.0
    bench_ann_vol = np.std(bench_ret, ddof=1) * np.sqrt(242) if len(bench_ret) > 1 else 0.0

    # 跟踪误差与信息比率
    if len(excess_ret) > 1 and np.std(excess_ret, ddof=1) > 0:
        tracking_err = np.std(excess_ret, ddof=1) * np.sqrt(242)
        info_ratio = np.mean(excess_ret) / np.std(excess_ret, ddof=1) * np.sqrt(242)
    else:
        tracking_err = 0.0
        info_ratio = 0.0

    # 相关系数
    corr = float(np.corrcoef(daily_ret, bench_ret)[0, 1]) if len(daily_ret) > 1 else 0.0

    # Beta
    if len(daily_ret) > 1 and np.var(bench_ret) > 0:
        beta = float(np.cov(daily_ret, bench_ret)[0, 1] / np.var(bench_ret))
    else:
        beta = 0.0

    # 8. 输出报告
    print("\n" + "=" * 70)
    print(f"  行业轮动策略 {version} OOS 净值分析 ({OOS_START} → {latest_date})")
    print("=" * 70)
    print(f"  交易日数:     {n_days}")
    print(f"  调仓次数:     {len(trade_dates)}")
    print()

    print("  ┌─────────────────┬──────────────┬──────────────┐")
    print("  │ 指标            │ 策略         │ 沪深300      │")
    print("  ├─────────────────┼──────────────┼──────────────┤")
    print(f"  │ 最终净值         │ {nav_arr[-1]:<12.4f} │ {bench_arr[-1]:<12.4f} │")
    print(f"  │ 累计收益         │ {total_return:>+11.2%} │ {bench_total_return:>+11.2%} │")
    print(f"  │ 年化Sharpe       │ {sharpe:<12.4f} │ {bench_sharpe:<12.4f} │")
    print(f"  │ 年化波动率       │ {ann_vol:<12.2%} │ {bench_ann_vol:<12.2%} │")
    print(f"  │ 最大回撤         │ {max_drawdown:>+11.2%} │ {bench_max_dd:>+11.2%} │")
    print("  └─────────────────┴──────────────┴──────────────┘")
    print()
    print(f"  超额收益(累计):   {excess_total:>+.2%}")
    print(f"  跟踪误差(年化):   {tracking_err:.2%}")
    print(f"  信息比率(IR):     {info_ratio:.4f}")
    print(f"  相关系数:         {corr:.4f}")
    print(f"  Beta:             {beta:.4f}")

    # 9. 调仓日明细
    print("\n" + "─" * 70)
    print("  调仓日明细:")
    print("  ┌────────────┬───────────┬───────────┬──────────┬────────────┐")
    print("  │ 日期        │ 净值      │ 沪深300   │ 持仓数   │ 现金占比   │")
    print("  ├────────────┼───────────┼───────────┼──────────┼────────────┤")
    for d in trade_dates:
        row = nav_df.filter(pl.col("date") == d)
        if len(row) > 0:
            r = row.to_dicts()[0]
            cash_ratio = r["cash"] / (r["cash"] + r["market_value"]) if (r["cash"] + r["market_value"]) > 0 else 0
            print(f"  │ {d} │ {r['nav']:<9.4f} │ {r['bench_nav']:<9.4f} │ {r['n_holdings']:<8d} │ {cash_ratio:<10.2%} │")
    print("  └────────────┴───────────┴───────────┴──────────┴────────────┘")

    # 10. 逐月统计
    print("\n" + "─" * 70)
    print("  逐月表现:")
    nav_df_monthly = nav_df.with_columns(
        pl.col("date").str.slice(0, 7).alias("month")
    )
    monthly = nav_df_monthly.group_by("month").agg([
        pl.col("nav").first().alias("nav_start"),
        pl.col("nav").last().alias("nav_end"),
        pl.col("bench_nav").first().alias("bench_start"),
        pl.col("bench_nav").last().alias("bench_end"),
    ]).sort("month")

    print("  ┌─────────┬──────────────┬──────────────┬──────────────┐")
    print("  │ 月份    │ 策略月收益   │ 沪深300月收益│ 超额月收益   │")
    print("  ├─────────┼──────────────┼──────────────┼──────────────┤")
    for r in monthly.to_dicts():
        s_ret = r["nav_end"] / r["nav_start"] - 1
        b_ret = r["bench_end"] / r["bench_start"] - 1
        e_ret = s_ret - b_ret
        print(f"  │ {r['month']} │ {s_ret:>+12.2%} │ {b_ret:>+12.2%} │ {e_ret:>+12.2%} │")
    print("  └─────────┴──────────────┴──────────────┴──────────────┘")

    # 11. 交易成本分析
    print("\n" + "─" * 70)
    total_cost = 0.0
    total_turnover = 0.0
    for f in files:
        trades = load_trades(f)
        for t in trades:
            total_cost += t["费用"]
            total_turnover += t["成交金额"]

    print(f"  累计交易金额:    {total_turnover:>14,.0f}")
    print(f"  累计交易成本:    {total_cost:>14,.0f}")
    print(f"  成本占初始资金:  {total_cost / CAPITAL:>14.2%}")
    print(f"  成本占交易额:    {total_cost / total_turnover if total_turnover > 0 else 0:>14.4f}")
    print(f"  平均单次调仓成本: {total_cost / len(files):>14,.0f}")

    # 12. 优化建议
    print("\n" + "─" * 70)
    print("  优化建议分析:")
    print("─" * 70)

    # 现金占用分析
    avg_cash_ratio = float((nav_df["cash"] / (nav_df["cash"] + nav_df["market_value"])).mean())
    print(f"  [1] 现金占用: 平均现金占比 {avg_cash_ratio:.2%}")

    # 诊断现金来源：建仓时实际买入金额 vs 目标仓位
    build_file = files[0]  # 20260601_build.xlsx
    build_trades = load_trades(build_file)
    build_buy_amount = sum(t["成交金额"] for t in build_trades if t["业务类型"] == "买入")
    build_position_ratio = build_buy_amount / CAPITAL
    print(f"      建仓实际买入金额: {build_buy_amount:,.0f} (占初始资金 {build_position_ratio:.2%})")
    if build_position_ratio < 0.85:
        print(f"      ⚠ 建仓仓位偏低，来源诊断:")
        print(f"        - 可能是 regime_adaptive 风控减仓 (min_exposure_scale={0.3})")
        print(f"        - 或大盘趋势过滤触发 (market_ma_short < market_ma_long)")
        print(f"        建议: 检查建仓日大盘状态，确认是否为风控主动减仓")

    if avg_cash_ratio > 0.05 and build_position_ratio > 0.85:
        # 建仓仓位满但后续现金占比高 → 整手取整问题
        print(f"      ⚠ 后续现金占比偏高，建议优化整手取整逻辑（LOT_SIZE=100）")
    elif avg_cash_ratio > 0.05:
        print(f"      注: 现金占比偏高主要来自风控减仓，属策略设计（下跌市减仓避险）")

    # 换手率分析
    # 单边换手率 = 成交金额 / 平均总资产
    avg_asset = float((nav_df["cash"] + nav_df["market_value"]).mean())
    single_side_turnover = total_turnover / 2 / avg_asset
    annual_turnover = single_side_turnover * (242 / n_days) if n_days > 0 else 0
    print(f"\n  [2] 换手率: 单边累计 {single_side_turnover:.2%}, 年化约 {annual_turnover:.2%}")
    if annual_turnover > 10:
        print(f"      ⚠ 换手率偏高（周频调仓导致），交易成本侵蚀收益")
        print(f"        - 考虑加入调仓阈值：仅当权重偏离超过 X% 时才调仓")
        print(f"        - 或改用双周/月频调仓")
    else:
        print(f"      ✓ 换手率可控")

    # 超额收益分析
    print(f"\n  [3] 超额收益: 累计 {excess_total:+.2%}, IR={info_ratio:.4f}")
    if excess_total > 0:
        print(f"      ✓ 策略跑赢沪深300")
    else:
        print(f"      ⚠ 策略跑输沪深300，需检查:")
        print(f"        - 行业轮动信号是否失效")
        print(f"        - 大盘过滤是否过于敏感")

    # 回撤分析
    print(f"\n  [4] 回撤控制: 策略最大回撤 {max_drawdown:.2%} vs 沪深300 {bench_max_dd:.2%}")
    # 注: 回撤为负数，数值越大（越接近0）表示回撤越小越优
    if max_drawdown > bench_max_dd:
        print(f"      ✓ 回撤控制优于沪深300")
    else:
        print(f"      ⚠ 回撤大于沪深300，建议:")
        print(f"        - 加强 regime_adaptive 风控的敏感度")
        print(f"        - 或增加止损逻辑")

    # 持仓集中度分析
    print(f"\n  [5] 持仓集中度:")
    # 取最新持仓
    last_date = trade_dates[-1]
    last_holdings = holdings_by_date[last_date]
    total_shares_value = 0
    code_values = []
    # 用最新日收盘价估值
    last_row = price_pivot.filter(pl.col("date") == latest_date)
    if len(last_row) > 0:
        last_row_dict = last_row.to_dicts()[0]
        for code, shares in last_holdings.items():
            close = last_row_dict.get(code)
            if close and not np.isnan(close):
                val = shares * close
                code_values.append((code, val))
                total_shares_value += val
        code_values.sort(key=lambda x: -x[1])
        print(f"      最新持仓 ({latest_date} 估值):")
        for code, val in code_values:
            w = val / total_shares_value if total_shares_value > 0 else 0
            print(f"        {code:<12} {w:>7.2%}")
        # HHI 指数
        weights = [v / total_shares_value for _, v in code_values if total_shares_value > 0]
        hhi = sum(w * w for w in weights)
        print(f"      HHI 集中度: {hhi:.4f} (1/N={1/max(len(weights),1):.4f})")
        if hhi > 0.2:
            print(f"      ⚠ 持仓集中度偏高，单股权重过大")

    # 相关性分析
    print(f"\n  [6] 与沪深300相关性: {corr:.4f}, Beta={beta:.4f}")
    if corr > 0.9:
        print(f"      ⚠ 与沪深300高度相关，分散化不足")
    elif corr < 0.3:
        print(f"      ✓ 与沪深300相关性低，具有分散化价值")

    print("\n" + "=" * 70)
    print("  分析完成")
    print("=" * 70)


if __name__ == "__main__":
    main()
