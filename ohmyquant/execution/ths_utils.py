"""同花顺 PMS 交易流水文件生成工具

提供跨策略通用的同花顺交易文件生成功能,避免 scripts/ 下脚本跨策略 import。
被 scripts/industry_rotation_daily.py 和 scripts/expertforest_v1_position_analysis.py
等建仓/调仓脚本复用。

核心函数:
  - get_open_prices: 获取指定日期开盘价(实际市场价格,停牌时用最近交易日收盘价替代)
  - compute_lot_shares: 按资金权重和价格计算整手股数
  - generate_trades: 生成建仓/调仓交易流水
  - write_xlsx: 写入同花顺 PMS 模板 xlsx
  - replay_history: 回放历史调仓重建持仓状态

用法:
  from ohmyquant.execution.ths_utils import (
      get_open_prices, generate_trades, write_xlsx, replay_history,
      CAPITAL, TRANSACTION_COST_RATE,
  )
"""
from __future__ import annotations

import logging
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from ohmyquant.data.sources.duckdb_source import DuckDBSource

logger = logging.getLogger(__name__)

# 默认交易参数(可被调用方覆盖)
CAPITAL = 10_000_000
TRANSACTION_COST_RATE = 0.001
LOT_SIZE = 100
TEMPLATE_PATH = Path("templates/ths_pms_template.xlsx")


def get_open_prices(source: DuckDBSource, codes: list[str], date_str: str) -> dict[str, float]:
    """获取指定日期开盘价(实际市场价格,非复权),缺失时用最近交易日收盘价替代

    同花顺 PMS 需要实际市场价格成交,不能用后复权价格。
    停牌/退市等导致当日无开盘价时,用最近交易日收盘价替代,避免漏买漏卖。
    """
    # 用 adjust="none" 获取实际市场价格(非复权),同花顺按实际价成交
    df = source.load_daily_price(codes, date_str, date_str, adjust="none")

    result: dict[str, float] = {}
    if df is not None and len(df) > 0:
        for row in df.iter_rows(named=True):
            code = row.get("code", "")
            open_price = row.get("open")
            if open_price and isinstance(open_price, (int, float)) and open_price > 0:
                result[code] = float(open_price)

    # 对缺失开盘价的股票,向前查找最近交易日的收盘价替代
    missing_codes = [c for c in codes if c not in result]
    if missing_codes:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        # 向前查找最多 10 个自然日,覆盖周末和短假期
        for days_back in range(1, 11):
            still_missing = [c for c in missing_codes if c not in result]
            if not still_missing:
                break
            prev_date = (dt - timedelta(days=days_back)).strftime("%Y-%m-%d")
            prev_df = source.load_daily_price(still_missing, prev_date, prev_date, adjust="none")
            if prev_df is not None and len(prev_df) > 0:
                for row in prev_df.iter_rows(named=True):
                    code = row.get("code", "")
                    close_price = row.get("close")
                    if close_price and isinstance(close_price, (int, float)) and close_price > 0:
                        result[code] = float(close_price)
                        logger.debug(f"{date_str} {code}: 开盘价缺失,用 {prev_date} 收盘价 {close_price} 替代")

    return result


def compute_lot_shares(capital: float, weight: float, price: float,
                       lot_size: int = LOT_SIZE) -> int:
    """按资金权重和价格计算整手股数(A股100股一手)"""
    if price <= 0:
        return 0
    raw_shares = capital * weight / price
    return int(raw_shares // lot_size) * lot_size


def generate_trades(
    date_str: str,
    prev_shares: dict[str, int],
    target_holdings: dict[str, float],
    open_prices: dict[str, float],
    prev_cash: float,
    is_build: bool,
    strategy_name: str = "strategy",
    capital: float = CAPITAL,
    cost_rate: float = TRANSACTION_COST_RATE,
    lot_size: int = LOT_SIZE,
) -> tuple[list[dict], dict[str, int], float]:
    """生成建仓/调仓交易流水

    Args:
        date_str: 交易日期 YYYY-MM-DD
        prev_shares: 上一调仓日持仓 {code: shares}
        target_holdings: 目标持仓权重 {code: weight}
        open_prices: 开盘价 {code: price}
        prev_cash: 上一调仓日后现金
        is_build: 是否为建仓(首次买入)
        strategy_name: 策略名称(用于交易说明)
        capital: 建仓初始资金(仅 is_build 时使用)
        cost_rate: 交易费率
        lot_size: 整手股数

    Returns:
        (trades, new_shares, new_cash)
    """
    trades: list[dict] = []
    dt = datetime.strptime(date_str, "%Y-%m-%d")

    if is_build:
        # 建仓:全部买入
        for code, weight in target_holdings.items():
            price = open_prices.get(code)
            if not price or price <= 0:
                continue
            shares = compute_lot_shares(capital, weight, price, lot_size)
            if shares <= 0:
                continue
            amount = shares * price
            cost = amount * cost_rate
            trades.append({
                "交易日期": dt, "证券代码": code, "业务类型": "买入",
                "数量": shares, "价格": round(price, 4),
                "成交金额": round(amount, 2), "费用": round(cost, 2),
                "证券类型": "A股", "说明": f"建仓 {strategy_name}",
            })
        new_shares = {t["证券代码"]: t["数量"] for t in trades}
        new_cash = prev_cash - sum(t["成交金额"] + t["费用"] for t in trades)
    else:
        # 调仓:先卖后买
        current_value = prev_cash
        for code, shares in prev_shares.items():
            current_value += shares * open_prices.get(code, 0)

        target_shares: dict[str, int] = {}
        for code, weight in target_holdings.items():
            price = open_prices.get(code)
            if not price or price <= 0:
                continue
            target_shares[code] = compute_lot_shares(current_value, weight, price, lot_size)

        # 卖出
        for code, shares in prev_shares.items():
            price = open_prices.get(code, 0)
            if price <= 0:
                continue
            target = target_shares.get(code, 0)
            if shares > target:
                sell_shares = shares - target
                amount = sell_shares * price
                cost = amount * cost_rate
                trades.append({
                    "交易日期": dt, "证券代码": code, "业务类型": "卖出",
                    "数量": sell_shares, "价格": round(price, 4),
                    "成交金额": round(amount, 2), "费用": round(cost, 2),
                    "证券类型": "A股", "说明": f"调仓卖出 {strategy_name}",
                })

        # 买入
        for code, target in target_shares.items():
            price = open_prices.get(code, 0)
            if price <= 0:
                continue
            current = prev_shares.get(code, 0)
            if target > current:
                buy_shares = target - current
                amount = buy_shares * price
                cost = amount * cost_rate
                trades.append({
                    "交易日期": dt, "证券代码": code, "业务类型": "买入",
                    "数量": buy_shares, "价格": round(price, 4),
                    "成交金额": round(amount, 2), "费用": round(cost, 2),
                    "证券类型": "A股", "说明": f"调仓买入 {strategy_name}",
                })

        new_shares = {c: target_shares.get(c, 0) for c in set(list(prev_shares.keys()) + list(target_shares.keys()))}
        new_shares = {c: s for c, s in new_shares.items() if s > 0}
        new_cash = prev_cash
        for t in trades:
            if t["业务类型"] == "买入":
                new_cash -= t["成交金额"] + t["费用"]
            else:
                new_cash += t["成交金额"] - t["费用"]

    return trades, new_shares, new_cash


def write_xlsx(trades: list[dict], output_path: Path,
               template_path: Path = TEMPLATE_PATH) -> None:
    """写入同花顺 PMS 模板 xlsx

    Args:
        trades: 交易流水列表
        output_path: 输出文件路径
        template_path: 同花顺模板路径
    """
    wb = load_workbook(template_path)
    ws = wb["Sheet1"]
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None
    headers = ["交易日期", "证券代码", "业务类型", "数量", "价格",
               "成交金额", "费用", "证券类型", "说明"]
    for i, trade in enumerate(trades, start=2):
        for j, key in enumerate(headers, start=1):
            ws.cell(i, j).value = trade[key]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def replay_history(
    source: DuckDBSource,
    rebalance_log: list[dict],
    check_date: str,
    strategy_name: str = "strategy",
    capital: float = CAPITAL,
    cost_rate: float = TRANSACTION_COST_RATE,
    lot_size: int = LOT_SIZE,
) -> tuple[dict[str, int], float]:
    """回放历史调仓,重建到 check_date 前一日的持仓状态

    遍历 check_date 之前的所有调仓日,回放交易计算持仓和现金。
    无需依赖 state.json,每次从OOS回测结果完整重建。

    Args:
        source: DuckDBSource 数据源
        rebalance_log: 调仓日志 [{date, holdings, ...}]
        check_date: 检查日期 YYYY-MM-DD
        strategy_name: 策略名称
        capital: 建仓初始资金
        cost_rate: 交易费率
        lot_size: 整手股数

    Returns:
        (prev_shares, prev_cash) check_date 前一日的持仓和现金
    """
    prev_shares: dict[str, int] = {}
    prev_cash = capital

    for entry in rebalance_log:
        date_str = entry["date"]
        if date_str >= check_date:
            break  # 只回放 check_date 之前的

        holdings = entry["holdings"]
        # 获取所有股票的开盘价(当前持仓 + 目标持仓),避免漏卖漏买
        all_codes = list(set(list(prev_shares.keys()) + list(holdings.keys())))
        open_prices = get_open_prices(source, all_codes, date_str)

        is_build = (len(prev_shares) == 0)
        _, prev_shares, prev_cash = generate_trades(
            date_str, prev_shares, holdings, open_prices, prev_cash, is_build,
            strategy_name=strategy_name,
            capital=capital,
            cost_rate=cost_rate,
            lot_size=lot_size,
        )

    return prev_shares, prev_cash


__all__ = [
    # 常量
    "CAPITAL",
    "TRANSACTION_COST_RATE",
    "LOT_SIZE",
    "TEMPLATE_PATH",
    # 函数
    "get_open_prices",
    "compute_lot_shares",
    "generate_trades",
    "write_xlsx",
    "replay_history",
]
